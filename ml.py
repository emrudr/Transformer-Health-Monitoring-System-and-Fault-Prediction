{\rtf1\ansi\ansicpg1252\cocoartf2761
\cocoatextscaling0\cocoaplatform0{\fonttbl\f0\fnil\fcharset0 HelveticaNeue;\f1\fnil\fcharset0 AppleColorEmoji;}
{\colortbl;\red255\green255\blue255;\red0\green0\blue0;}
{\*\expandedcolortbl;;\cssrgb\c0\c0\c0;}
\paperw11900\paperh16840\margl1440\margr1440\vieww11520\viewh8400\viewkind0
\pard\tx560\tx1120\tx1680\tx2240\tx2800\tx3360\tx3920\tx4480\tx5040\tx5600\tx6160\tx6720\pardirnatural\partightenfactor0

\f0\fs24 \cf2 import serial\
import pandas as pd\
from openpyxl import load_workbook\
import time\
import re\
from sklearn.preprocessing import StandardScaler\
from sklearn.ensemble import IsolationForest\
from sklearn.metrics.pairwise import cosine_similarity\
# Serial port configuration\
SERIAL_PORT = "/dev/cu.usbserial-10"\
BAUD_RATE = 115200\
OUTPUT_FILE = "serial_data.xlsx"\
# Define column headers\
columns = ["Timestamp", "Current (A)", "Temperature (\'b0C)", "Humidity (%)", "Voltage (V)",\
"Vibration"]\
# Try to load existing Excel file or create a new one\
try:\
workbook = load_workbook(OUTPUT_FILE)\
sheet = workbook.active\
except FileNotFoundError:\
df = pd.DataFrame(columns=columns)\
df.to_excel(OUTPUT_FILE, index=False)\
workbook = load_workbook(OUTPUT_FILE)\
sheet = workbook.active\
print(f"Connected to \{SERIAL_PORT\} at \{BAUD_RATE\} baud. Logging data to Excel...\
Press CTRL+C to stop.")\
# Initialize Serial Connection\
try:\
ser = serial.Serial(SERIAL_PORT, BAUD_RATE, timeout=1)\
38except serial.SerialException:\
print(f"Error: Could not connect to \{SERIAL_PORT\}")\
exit()\
# Function to extract values from serial data\
def parse_serial_data(serial_lines):\
data_dict = \{\}\
for line in serial_lines:\
match = re.match(r"(Current|Temperature|Humidity|Voltage|Vibration):\\s*([-+]?\\d*\\.?\
\\d+)", line)\
if match:\
key = match.group(1) # Extract the label (e.g., "Current")\
value = match.group(2) # Extract the numeric value\
data_dict[key] = float(value) # Convert to float for Excel storage\
# Ensure all values exist, otherwise assign None (to maintain alignment)\
return [\
time.strftime("%Y-%m-%d %H:%M:%S"), # Timestamp\
data_dict.get("Current", None),\
data_dict.get("Temperature", None),\
data_dict.get("Humidity", None),\
data_dict.get("Voltage", None),\
data_dict.get("Vibration", None),\
]\
# ========== ML Model Setup ==========\
def train_model(df):\
# Train Isolation Forest on initial data\
features = ['Current (A)', 'Temperature (\'b0C)', 'Humidity (%)', 'Voltage (V)', 'Vibration']\
X_train = df[features]\
scaler = StandardScaler()\
X_scaled = scaler.fit_transform(X_train)\
model = IsolationForest(contamination=0.01, random_state=42)\
39model.fit(X_scaled)\
# Save shutdown pattern as last row (can change for more robust detection)\
shutdown_pattern = scaler.transform([X_train.iloc[-1].values])\
return model, scaler, shutdown_pattern\
# ========== Initialize model with initial data ==========\
try:\
# Wait until we have enough data to train the model (50 rows)\
while len(sheet['A']) < 51:\
time.sleep(2)\
# Load initial data for model training\
df_initial = pd.read_excel(OUTPUT_FILE)\
model, scaler, shutdown_pattern = train_model(df_initial)\
print("
\f1 \uc0\u9989  
\f0 Model trained. Starting live monitoring...")\
except Exception as e:\
print(f"Error during model training: \{e\}")\
exit()\
# ========== Data Logging Loop ==========\
final_report = []\
try:\
while True:\
if ser.in_waiting > 0:\
serial_data = []\
# Read multiple lines (as each data entry is on a separate line)\
while ser.in_waiting:\
line = ser.readline().decode("utf-8", errors="ignore").strip()\
if line:\
serial_data.append(line)\
40parsed_values = parse_serial_data(serial_data)\
if parsed_values:\
# Log data to Excel\
sheet.append(parsed_values)\
workbook.save(OUTPUT_FILE)\
# Process the new data for anomaly detection\
df_new = pd.read_excel(OUTPUT_FILE)\
new_data = df_new.iloc[-1][\
['Current (A)', 'Temperature (\'b0C)', 'Humidity (%)', 'Voltage (V)',\
'Vibration']].values\
input_scaled = scaler.transform([new_data])\
prediction = model.predict(input_scaled)[0]\
entry = \{\
"Timestamp": parsed_values[0],\
"Data": parsed_values[1:],\
"Status": "",\
"Similarity": None\
\}\
# Check if the prediction is an anomaly\
if prediction == -1:\
similarity = cosine_similarity(input_scaled, shutdown_pattern)[0][0]\
entry["Similarity"] = round(similarity, 2)\
if similarity > 0.95:\
print(f"\\n
\f1 \uc0\u55357 \u57041  
\f0 TRIP SIGNAL! Pattern matched for shutdown event at\
\{parsed_values[0]\}.")\
entry["Status"] = "Trip Triggered"\
else:\
print(f"\\n
\f1 \uc0\u9888  
\f0 Anomaly Detected at \{parsed_values[0]\} (not critical).")\
entry["Status"] = "Anomaly"\
else:\
print(f"
\f1 \uc0\u9989  
\f0 Normal reading at \{parsed_values[0]\}.")\
entry["Status"] = "Normal"\
41final_report.append(entry)\
except KeyboardInterrupt:\
print("\\nLogging stopped by user.")\
ser.close()\
workbook.save(OUTPUT_FILE)\
print(f"Data saved to \{OUTPUT_FILE\}")\
# Save final report\
report_df = pd.DataFrame(final_report)\
report_df[['Current (A)', 'Temperature (\'b0C)', 'Humidity (%)', 'Voltage (V)', 'Vibration']] =\
pd.DataFrame(\
report_df['Data'].tolist(), index=report_df.index)\
report_df.drop(columns='Data', inplace=True)\
report_df.to_excel("final_report.xlsx", index=False)\
print("
\f1 \uc0\u9989  
\f0 Final report saved as 'final_report.xlsx'")}